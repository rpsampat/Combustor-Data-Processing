import math
import os
import cantera as ct
import numpy as np
import pickle
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from mpl_toolkits.axes_grid1 import host_subplot
from mpl_toolkits import axisartist
#import Thermocouple as TC
from labview_extract import DataExtract
from scandir import scandir
import Settings as st

class Profiles_intercase:
    def __init__(self):
        self.drive = 'P:/'
        self.folder = "GasComposition and Temperature_April2022/"
        self.dataset={}
        self.dataset_temperature={}
        self.quant_list = ["Measured - Main Air", "Measured - Pilot Air", "Measured - Cooling Air",
                           "Measured - CH4 Main", "'Raw TC Data'/'TC15'", "'Raw TC Data'/'Main Air'",
                           "'Raw TC Data'/'TC10'",
                           "'Raw TC Data'/'Cooling Air'", "'Raw TC Data'/'Fuel Temp'", "Offset",
                           "Position", "Avg Exhaust Temp", "'Pressure & GA'/'CH4'", "'Pressure & GA'/'CO2'",
                           "'Pressure & GA'/'CO'", "'Pressure & GA'/'NO'", "'Pressure & GA'/'NO2'",
                           "'Pressure & GA'/'O2'"]
        self.quant = {'mdot_air': "Measured - Main Air", 'mdot_pilotair': "Measured - Pilot Air",
                      'mdot_coolair': "Measured - Cooling Air", 'mdot_ch4': "Measured - CH4 Main",
                      'TC15': "'Raw TC Data'/'TC15'", 'Temp_mainair': "'Raw TC Data'/'Main Air'",
                      'Temp_flange': "'Raw TC Data'/'TC10'", 'Temp_coolair': "'Raw TC Data'/'Cooling Air'",
                      'Temp_fuel': "'Raw TC Data'/'Fuel Temp'", 'y_pos': "Offset", 'x_pos': "Position",
                      'T_exhaust': "Avg Exhaust Temp", 'CH4': "'Pressure & GA'/'CH4'", 'CO2': "'Pressure & GA'/'CO2'",
                      'CO': "'Pressure & GA'/'CO'", 'NO': "'Pressure & GA'/'NO'", 'NO2': "'Pressure & GA'/'NO2'",
                      'O2': "'Pressure & GA'/'O2'"}
        self.diluent_list = ["None", "N2","CO2"]
        self.dilution_list = ["20.96", "19", "17.5", "14"]
        self.power_list = [30, 40, 50, 60]
        self.phi_list = [0.9, 0.8, 0.7, 0.6]

    def search_files(self,power,diluent,dilution,phi):
        directory = self.drive+self.folder
        subdir_list = [x[0] for x in os.walk(directory)]
        if diluent=="None":
            filename = "/gasprobe_temp_" + str(power) + "kW_phi" + str(phi)+ ".pkl"
        else:
            filename = "/gasprobe_temp_" + str(power) + "kW_phi" + str(phi)+"_" + diluent + "_" + str(dilution) + "perc.pkl"

        for subdir in subdir_list:
            try:
                with open(subdir + filename, 'rb') as f:
                    data = pickle.load(f)
                return data
            except:
                continue

        return []

    def unified_dataset(self):

        diluent_list = ["None","N2", "CO2"]
        dilution_list = ["20.96","19", "17.5", "14"]
        power_list = [30, 40, 50, 60]
        phi_list = [0.9, 0.8, 0.7, 0.6]
        for power in power_list:
            self.dataset[power] = {}
            for phi in phi_list:
                self.dataset[power][phi]={}
                for dilution in dilution_list:
                    self.dataset[power][phi][dilution] = {}
                    for diluent in diluent_list:
                        if dilution == "20.96" and diluent!="None":
                            continue
                        self.dataset[power][phi][dilution][diluent] = self.search_files(power,diluent,dilution,phi)

        with open(self.drive+self.folder + "Unified_dataset_gascomp", 'wb') as file:
            pickle.dump(self.dataset, file, pickle.HIGHEST_PROTOCOL)

    def exhaust_unified_dataset(self):

        diluent_list = ["None"]#,"N2", "CO2"]
        dilution_list = ["20.96"]#,"19", "17.5", "14"]
        power_list = [30]#, 40, 50, 60]
        phi_list = [0.95,0.9, 0.8, 0.7, 0.6]
        for power in power_list:
            self.dataset[power] = {}
            for phi in phi_list:
                self.dataset[power][phi]={}
                for dilution in dilution_list:
                    self.dataset[power][phi][dilution] = {}
                    for diluent in diluent_list:
                        if dilution == "20.96" and diluent!="None":
                            continue
                        self.dataset[power][phi][dilution][diluent] = self.search_files(power,diluent,dilution,phi)

        with open(self.drive+self.folder + "Unified_dataset_gascomp", 'wb') as file:
            pickle.dump(self.dataset, file, pickle.HIGHEST_PROTOCOL)

    def sweep_zaber(self, data, t_pretrigger, pos_id,quant_id,quant_label,freq_acq):
        """
        Several x locations bundled together in one acquisition
        :param data:
        :param t_pretrigger:
        :param pos_id:'x_pos'/'y_pos'
        :param quant_id:'TC15'/'NO'...
        :return:
        """
        xlist = []
        Z_mean = {}
        Z_stdv = {}
        Z_uncert ={}
        Z_max = {}
        pos =0
        t_acquire = 200  # s

        for bunch in data.keys():
            xpos_list = np.repeat(data[bunch][self.quant[pos_id]], freq_acq)
            #ypos_list = np.repeat(data[bunch][self.quant['y_pos']], 1)# temperature logged at 10Hz and positions at 1Hz
            unq_x, unq_ind_x, unq_freq_x = np.unique(xpos_list, return_counts=True, return_index=True)
            freq_min = 2000
            ind_steady_x = [i for i in range(len(unq_x)) if (unq_freq_x[i] > freq_min )] # checking for unique values with more than freq_min occurrences
            Z_mean[bunch] = []
            Z_stdv[bunch] = []
            Z_uncert[bunch] = []
            Z_max[bunch]=[]
            x_plot =[]
            for x_loc in ind_steady_x:
                pos +=1
                xlist.append((xpos_list[
                              unq_ind_x[x_loc] + t_pretrigger * freq_acq]))
                temp_list = data[bunch][self.quant[quant_id]][
                            unq_ind_x[x_loc] + t_pretrigger * freq_acq:unq_ind_x[x_loc]+t_acquire*freq_acq]#+unq_freq_x[x_loc]]
                Z_mean[bunch].append(np.mean(temp_list))
                Z_stdv[bunch].append(np.std(temp_list))
                Z_uncert[bunch].append(np.std(temp_list)/np.sqrt(len(temp_list)))
                x_plot.append(unq_x[x_loc])

        return Z_mean[bunch], Z_uncert[bunch], x_plot

    def singlept_dataacq_extract(self, data, quant_id):
        xlist = []
        Z_mean = {}
        Z_stdv = {}
        Z_uncert = {}
        Z_max = {}
        pos = 0

        for bunch in data.keys():
            temp_list = data[bunch][self.quant[quant_id]]
            Z_mean[bunch]=[(np.mean(temp_list))]
            Z_stdv[bunch]=[(np.std(temp_list))]
            Z_uncert[bunch]=[(np.std(temp_list) / np.sqrt(len(temp_list)))]

        return Z_mean[bunch], Z_uncert[bunch]

    def sweep_zaber_single(self, data, t_pretrigger, pos_id,quant_id,quant_label,freq_acq):
        """
        Slow traversal
        :param data:
        :param t_pretrigger:
        :param pos_id:'x_pos'/'y_pos'
        :param quant_id:'TC15'/'NO'...
        :return:
        """
        x_legend = []
        xlist = []
        Z_mean = {}
        Z_stdv = {}
        Z_uncert ={}
        #fig, ax = plt.subplots()
        #fig2, ax2 = plt.subplots()
        plotid = 0
        pos =0
        t_acquire = 200 #s
        for bunch in data.keys():
            xpos_list = np.repeat(data[bunch][self.quant[pos_id]], freq_acq)
            #ypos_list = np.repeat(data[bunch][self.quant['y_pos']], 1)# temperature logged at 10Hz and positions at 1Hz
            unq_x, unq_ind_x, unq_freq_x = np.unique(xpos_list, return_counts=True, return_index=True)
            freq_min = 2
            ind_steady_x = [i for i in range(len(unq_x)) if (unq_freq_x[i] > freq_min )] # checking for unique values with more than freq_min occurrences
            Z_mean[bunch] = []
            Z_stdv[bunch] = []
            Z_uncert[bunch] = []
            x_plot =[]
            for x_loc in ind_steady_x:
                pos +=1
                xlist.append((xpos_list[
                              unq_ind_x[x_loc]]))
                temp_list = data[bunch][self.quant[quant_id]][
                            unq_ind_x[x_loc]:unq_ind_x[x_loc]+unq_freq_x[x_loc]]
                Z_mean[bunch].append(np.mean(temp_list))
                Z_stdv[bunch].append(np.std(temp_list))
                Z_uncert[bunch].append(np.std(temp_list)/np.sqrt(len(temp_list)))

                x_plot.append(unq_x[x_loc])
            x_legend.append(str(self.phi_list[bunch-1]))

        return Z_mean[bunch],Z_uncert[bunch],x_plot

    def sweep_zaber_slowtrav_with_pointwise(self, data, t_pretrigger, pos_id,quant_id,quant_label,freq_acq):
        """
        Temperature data acquired by slow traversal but stored in same file as those with point wise measurements.
        It is assumed that the slow traverse data is at the end of the file.
        :param data:
        :param t_pretrigger:
        :param pos_id:'x_pos'/'y_pos'
        :param quant_id:'TC15'/'NO'...
        :return:
        """
        x_legend = []
        xlist = []
        x_pointwise_list=[]
        Z_mean = {}
        Z_stdv = {}
        Z_uncert ={}
        #fig, ax = plt.subplots()
        #fig2, ax2 = plt.subplots()
        plotid = 0
        pos =0
        t_acquire = 200 #s
        for bunch in data.keys():
            xpos_list = np.repeat(data[bunch][self.quant[pos_id]], freq_acq)
            #ypos_list = np.repeat(data[bunch][self.quant['y_pos']], 1)# temperature logged at 10Hz and positions at 1Hz
            unq_x, unq_ind_x, unq_freq_x = np.unique(xpos_list, return_counts=True, return_index=True)
            freq_min = 2000
            ind_steady_x_point = [i for i in range(len(unq_x)) if (unq_freq_x[i] > freq_min )] # checking for unique values with more than freq_min occurrences
            for x_loc in ind_steady_x_point:
                pos +=1
                x_pointwise_list.append((xpos_list[
                              unq_ind_x[x_loc] + t_pretrigger * freq_acq]))
            Z_mean[bunch] = []
            Z_stdv[bunch] = []
            Z_uncert[bunch] = []
            x_plot =[]
            freq_min = 2
            freq_max = 25
            ind_steady_x = [i for i in range(len(unq_x)) if (
                        unq_freq_x[i] > freq_min and unq_freq_x[i] < freq_max and unq_ind_x[i]>unq_ind_x[ind_steady_x_point[-1]])]  # checking for unique values with more than freq_min occurrences
            for x_loc in ind_steady_x:
                pos +=1
                xlist.append((xpos_list[
                              unq_ind_x[x_loc]]))

                temp_list = data[bunch][self.quant[quant_id]][
                            unq_ind_x[x_loc] :unq_ind_x[x_loc]+unq_freq_x[x_loc]]
                Z_mean[bunch].append(np.mean(temp_list))
                Z_stdv[bunch].append(np.std(temp_list))
                Z_uncert[bunch].append(np.std(temp_list)/np.sqrt(len(temp_list)))

                x_plot.append(unq_x[x_loc])
            x_legend.append(str(self.phi_list[bunch-1]))

        return Z_mean[bunch],Z_uncert[bunch],x_plot


    def emissions_vs_radialoc(self,power,phi,diluent,dilution):
        t_pretrigger = 30
        quant_id = ['NO', 'NO2', 'CO', 'CO2', 'CH4', 'O2']
        quant_label = ['NO (ppm)', 'NO2 (ppm)', 'CO (ppm)', 'CO2 (vol%)', 'CH4 (ppm)', 'O2 (vol%)']
        freq_acq_GP = 1
        data = self.dataset[power][phi][dilution][diluent]
        for q in range(len(quant_id)):
            self.sweep_zaber_single(data, t_pretrigger, 'x_pos', quant_id[q], quant_label[q],
                                    freq_acq_GP)  # 'y_pos:TC, 'x_pos':GP

    def emissions_vs_equivalenceratio(self):
        t_pretrigger = 30
        quant_id = ['NO', 'NO2', 'CO', 'CO2', 'CH4', 'O2']
        quant_label = ['NO (ppm)', 'NO2 (ppm)', 'CO (ppm)', 'CO2 (vol%)', 'CH4 (ppm)', 'O2 (vol%)']
        freq_acq_GP = 1
        for q in range(len(quant_id)):
            fig,ax=plt.subplots()
            xlegend=[]
            for power in self.power_list:
                xlegend.append(str(power))
                value_list=[]
                for phi in self.phi_list:
                    dilution = "20.96"
                    diluent = "None"
                    data = self.dataset[power][phi][dilution][diluent]
                    data_mean = self.sweep_zaber(data, t_pretrigger, 'x_pos', quant_id[q], quant_label[q],
                                    freq_acq_GP)  # 'y_pos:TC, 'x_pos':GP
                    value_list.append(data_mean[1][-1])
                ax.scatter(self.phi_list,value_list)
            ax.legend(xlegend, title="kW")
            ax.set_ylabel(quant_id[q])
            ax.set_xlabel("Equivalence Ratio")
            plt.show()

    def temperature_vs_equivalenceratio(self,port):
        t_pretrigger = 30
        quant_id = ['TC15']
        quant_label = ['Temperature  (C)']
        freq_acq_GP = 10
        for q in range(len(quant_id)):
            fig,ax=plt.subplots()
            xlegend=[]
            for power in self.power_list:
                xlegend.append(str(power))
                value_list=[]
                for phi in self.phi_list:
                    dilution = 20.95
                    diluent = "None"
                    data = self.dataset_temperature[port][power][phi][dilution][diluent]
                    data_mean, data_uncert, x_plot = self.sweep_zaber(data, t_pretrigger, 'y_pos', quant_id[q], quant_label[q],
                                    freq_acq_GP)  # 'y_pos:TC, 'x_pos':GP
                    value_list.append(data_mean[-1])#last location
                ax.scatter(self.phi_list,value_list)
            ax.legend(xlegend, title="kW")
            ax.set_ylabel(quant_id[q])
            ax.set_xlabel("Equivalence Ratio")
            plt.show()

    def temperatures_vs_radialoc(self,port):
        t_pretrigger = 30
        quant_id = ['TC15']
        quant_label = ['Temperature  (C)']
        freq_acq_TC = 10
        self.diluent_list = ["None", "N2", "CO2"]
        self.dilution_list = [20.95, 19,17.5,14]
        self.power_list = [30, 40, 50, 60]
        self.phi_list = [0.9, 0.8, 0.7, 0.6]
        diluent = "None"

        #fig, ax = plt.subplots()
        #fig2, ax2 = plt.subplots()
        for q in range(len(quant_id)):
            for power in [40]:
                fig, ax = plt.subplots()
                x_legend = []
                for phi in self.phi_list:#for power in [40]:#self.power_list:
                    x_legend.append(phi)
                    dilution = 20.95
                    data = self.dataset_temperature[port][power][phi][dilution][diluent]
                    #Z_mean,Z_uncert,x_plot=self.sweep_zaber_single(data, t_pretrigger, 'y_pos', quant_id[q], quant_label[q],
                     #               freq_acq_TC)  # 'y_pos:TC, 'x_pos':GP
                    Z_mean, Z_uncert, x_plot = self.sweep_zaber(data, t_pretrigger, 'y_pos', quant_id[q],
                                                                       quant_label[q],
                                                                       freq_acq_TC)  # 'y_pos:TC, 'x_pos':GP
                    ax.scatter(x_plot, Z_mean)
                    #ax.errorbar(x_plot, Z_mean, yerr=Z_uncert, linestyle='', marker='o', capsize=6,
                               # markersize=3)
                    # ax2.scatter(x_plot, Z_stdv[bunch])
                ax.set_xlabel('Distance from wall (mm)')
                ax.set_ylabel('Temperature  (C)')
                ax.legend(x_legend)


        plt.show()

    def temperatures_vs_radialoc_slowtrav_vs_pointwise(self,port):
        self.drive = 'O:/'
        self.folder = "GasComposition and Temperature_April2022/Data plots/"
        t_pretrigger = 30
        quant_id = ['TC15']
        quant_label = ['Temperature  (C)']
        freq_acq_TC = 10
        self.diluent_list = ["None", "N2", "CO2"]
        self.dilution_list = [20.95, 19,17.5,14]
        self.power_list = [30,40]#, 50, 60]
        self.phi_list = [0.9, 0.8, 0.7, 0.6]
        diluent = "None"
        marker_list = ["*", "x", "x"]
        color_list = ['k', 'r', 'm', 'g']
        if port == 6:
            zaberid = 'x_pos'
        else:
            zaberid = 'y_pos'

        #fig, ax = plt.subplots()
        #fig2, ax2 = plt.subplots()
        for q in range(len(quant_id)):
            for power in self.power_list:
                fig, ax = plt.subplots(dpi=110)
                x_legend = []
                count =0
                for phi in self.phi_list:#for power in [40]:#self.power_list:

                    dilution = 20.95
                    if port==3:
                        # pointwise temperature measurements only taken at port 3
                        pointwise="Yes"
                        x_legend.append(str(phi)+" point wise")
                        data = self.dataset_temperature[port][pointwise][power][phi][dilution][diluent]
                        # @Port 3 40kW cases pointwise and continuous stored together; 30kW no continuous
                        Z_mean1,Z_uncert1,x_plot1=self.sweep_zaber(data, t_pretrigger, zaberid, quant_id[q], quant_label[q],
                                        freq_acq_TC)  # 'y_pos:TC, 'x_pos':GP
                        ax.scatter(x_plot1, Z_mean1, s=85, marker=marker_list[1], color=color_list[count])
                    if power == 40 and port==3:
                        Z_mean, Z_uncert, x_plot = self.sweep_zaber_slowtrav_with_pointwise(data, t_pretrigger, zaberid,
                                                                                            quant_id[q],
                                                                                            quant_label[q],
                                                                                            freq_acq_TC)  # 'y_pos:TC, 'x_pos':GP
                    else:
                        pointwise = "No"

                        data = self.dataset_temperature[port][pointwise][power][phi][dilution][diluent]
                        Z_mean, Z_uncert, x_plot = self.sweep_zaber_single(data, t_pretrigger, zaberid,
                                                                                            quant_id[q],
                                                                                            quant_label[q],
                                                                                            freq_acq_TC)  # 'y_pos:TC, 'x_pos':GP

                        ax.scatter(x_plot, Z_mean, s=5,marker=marker_list[0],color=color_list[count])
                        x_legend.append(str(phi))

                    count +=1
                    #ax.errorbar(x_plot, Z_mean, yerr=Z_uncert, linestyle='', marker='o', capsize=6,
                               # markersize=3)
                    # ax2.scatter(x_plot, Z_stdv[bunch])
                ax.set_xlabel('Distance from wall (mm)')
                ax.set_ylabel('Temperature  (C)')
                ax.legend(x_legend)
                fig_name = "Temperature_slowtrav_vs_pointwise_"+str(power)+"kW_port" + str(port)
                fig.savefig(self.drive + self.folder + fig_name + '.pdf')
                fig.savefig(self.drive + self.folder + fig_name + '.png')


        plt.show()

    def peaktemperatures_vs_equivalenceratio(self,port):
        self.drive = 'O:/'
        self.folder = "GasComposition and Temperature_April2022/Data plots/"
        t_pretrigger = 30
        #quant_id = ['TC15']
        #quant_label = ['Temperature  (C)']
        quant_id = ['TC15']
        quant_label = ['Temperature  (C)']
        freq_acq_TC = 10
        self.diluent_list = ["None", "N2", "CO2"]
        self.dilution_list = [20.95, 19,17.5,14]
        self.power_list = [30,40]#, 50, 60]
        self.phi_list = [0.9, 0.8, 0.7, 0.6]
        diluent = "None"
        marker_list = ["*", "x", "x"]
        color_list = ['k', 'r', 'm', 'g']
        if port == 6:
            zaberid = 'x_pos'
        else:
            zaberid = 'y_pos'

        #fig, ax = plt.subplots()
        #fig2, ax2 = plt.subplots()
        fig, ax = plt.subplots(dpi=110)
        x_legend = []
        count = 0
        for q in range(len(quant_id)):
            for power in self.power_list:
                Z_peak = []
                x_legend.append(power)
                for phi in self.phi_list:#for power in [40]:#self.power_list:
                    dilution = 20.95
                    try:
                        pointwise = "No"
                        data = self.dataset_temperature[port][pointwise][power][phi][dilution][diluent]
                        Z_mean, Z_uncert, x_plot = self.sweep_zaber_single(data, t_pretrigger, zaberid,
                                                                           quant_id[q],
                                                                           quant_label[q],
                                                                           freq_acq_TC)  # 'y_pos:TC, 'x_pos':GP
                        Z_peak.append(np.max(Z_mean))
                    except:
                        try:
                            pointwise = "Yes"
                            data = self.dataset_temperature[port][pointwise][power][phi][dilution][diluent]
                            try:
                                Z_mean, Z_uncert, x_plot = self.sweep_zaber_slowtrav_with_pointwise(data, t_pretrigger,
                                                                                                    zaberid,
                                                                                                    quant_id[q],
                                                                                                    quant_label[q],
                                                                                                    freq_acq_TC)  # 'y_pos:TC, 'x_pos':GP
                            except:
                                Z_mean, Z_uncert, x_plot = self.sweep_zaber(data, t_pretrigger, zaberid, quant_id[q],
                                                                               quant_label[q],
                                                                               freq_acq_TC)  # 'y_pos:TC, 'x_pos':GP
                            Z_peak.append(np.max(Z_mean))
                        except:
                            break

                ax.scatter(self.phi_list, Z_peak, s=30, marker=marker_list[1], color=color_list[count])
                count +=1
        ax.set_xlabel('Equivalence Ratio')
        ax.set_ylabel('Temperature  (C)')
        ax.legend(x_legend)
        fig_name = "PeakTemperature_vs_phi_Port" + str(port)
        fig.savefig(self.drive + self.folder + fig_name + '.pdf')
        fig.savefig(self.drive + self.folder + fig_name + '.png')


        plt.show()



    def emissions_vs_dilution(self):
        t_pretrigger = 30
        quant_id = ['NO', 'NO2', 'CO', 'CO2', 'CH4', 'O2']
        quant_label = ['NO (ppm)', 'NO2 (ppm)', 'CO (ppm)', 'CO2 (vol%)', 'CH4 (ppm)', 'O2 (vol%)']
        freq_acq_GP = 1
        for q in range(len(quant_id)):
            fig, ax = plt.subplots()
            xlegend = []
            for power in self.power_list:
                for diluent in self.diluent_list:
                    if diluent == "None":
                        continue
                    xlegend.append(str(power)+"_"+diluent)
                    xlist=[]
                    value_list = []
                    for dilution in self.dilution_list:
                        phi = 0.9
                        if dilution=="20.96":
                            data = self.dataset[power][phi][dilution]["None"]
                        else:
                            data = self.dataset[power][phi][dilution][diluent]
                        try:
                            data_mean = self.sweep_zaber(data, t_pretrigger, 'x_pos', quant_id[q], quant_label[q],
                                                         freq_acq_GP)  # 'y_pos:TC, 'x_pos':GP
                            value_list.append(data_mean[1][-1])
                            xlist.append(float(dilution))
                        except:
                            continue
                    ax.scatter(xlist, value_list)
            ax.legend(xlegend, title="kW_Diluent")
            ax.set_ylabel(quant_id[q])
            ax.set_xlabel("Dilution level")
            plt.show()

    def emission_correction(self, NO,NO2,CO,CO2,CH4,O2):
        """

        :param NO: ppm
        :param NO2: ppm
        :param CO: ppm
        :param CO2: vol %
        :param CH4: ppm
        :param O2: vol %
        :return:
        """
        X_NO_wet = NO*1e-6
        X_NO2_wet = NO2*1e-6
        X_CO_dry = CO*1e-6
        X_CO2_dry = CO2/100
        X_O2_dry = O2/100
        X_CH4_dry = CH4*1e-6
        X_N2_dry = 1-(X_O2_dry+X_CO2_dry+X_CO_dry+X_CH4_dry)
        perc_corr = 0.15
        X_NO_dry = X_NO_wet*(1+2*(X_CO2_dry+X_CO_dry))/((X_CO2_dry+X_CO_dry+X_CH4_dry+X_N2_dry)/(1-perc_corr))
        X_NO2_dry = X_NO2_wet * (1 + 2 * (X_CO2_dry + X_CO_dry)) / (
                    (X_CO2_dry + X_CO_dry + X_CH4_dry + X_N2_dry) / (1 - perc_corr))
        X_O2_wet = X_O2_dry/(3*X_CO2_dry+2*X_CO_dry+X_CH4_dry+X_N2_dry)
        X_CO_corr = X_CO_dry/((X_CO2_dry+X_CO_dry+X_CH4_dry+X_N2_dry)/(1-perc_corr))
        X_H2O = 1/(1+(1/(2*(X_CO2_dry+X_CO_dry))))

        return X_NO_dry,X_NO2_dry,X_O2_wet,X_CO_corr,X_H2O

    def excess_O2(self,mdot_ch4,mdot_air,CO,CO2,CH4,O2,H2O):
        X_CO_dry = CO * 1e-6
        X_CO2_dry = CO2 / 100.0
        X_O2_dry = O2 / 100.0
        X_CH4_dry = CH4 * 1e-6
        X_H2O = H2O
        MW_CH4 = 16.0
        MW_air = 28.8
        n_tot = (mdot_ch4/MW_CH4)/(X_CO_dry+X_CO2_dry+X_CH4_dry)
        n_O2 = (3*X_CO_dry+4*X_CO2_dry+2*X_O2_dry)*n_tot/2.0
        n_O2_in = mdot_air/(4.76*MW_air)
        excess_O2 = (n_O2-n_O2_in)/n_O2_in

        return excess_O2
    def correctedemissions_vs_equivalenceratio(self,path,port):

        t_pretrigger = 0
        quant_id = ['NO', 'NO2', 'CO', 'CO2', 'CH4', 'O2']
        quant_label = ['NO (ppm)', 'NO2 (ppm)', 'CO (ppm)', 'CO2 (vol%)', 'CH4 (ppm)', 'O2 (vol%)']
        extra_quant_id = ['mdot_air', 'mdot_pilotair', 'mdot_ch4']
        extra_quant_label = ['mdot_air', 'mdot_pilotair', 'mdot_ch4']
        freq_acq_GP = 1
        fig, ax = plt.subplots()
        fig2, ax2 = plt.subplots()
        fig3, ax3 = plt.subplots()
        fig4, ax4 = plt.subplots()
        fig5, ax5 = plt.subplots()
        xlegend = []
        marker_list = ["o", "x", "x"]
        color_list = ['k', 'r', 'm', 'g']
        count =-1
        for power in [30]:#,40,50,60]:
            xlegend.append(str(power))
            value_list = {}
            NOx_corr = []
            O2_wet_list=[]
            CO_corr_list=[]
            H2O_conc_list = []
            excess_O2_list=[]
            phi_corr_list=[]
            count+=1
            for phi in self.phi_list:
                dilution = "20.96"
                diluent = "None"
                data = self.dataset[power][phi][dilution][diluent]
                data1_mean, data1_uncert, x1_plot = self.sweep_zaber(data, t_pretrigger, 'x_pos', extra_quant_id[0],
                                                                     extra_quant_label[0],freq_acq_GP)
                data2_mean, data2_uncert, x2_plot = self.sweep_zaber(data, t_pretrigger, 'x_pos', extra_quant_id[1],
                                                                     extra_quant_label[1],freq_acq_GP)
                data3_mean, data3_uncert, x3_plot = self.sweep_zaber(data, t_pretrigger, 'x_pos', extra_quant_id[2],
                                                                     extra_quant_label[2],freq_acq_GP)
                mdot_air_tot = ((data1_mean[-1]+data2_mean[-1])*1.1973/60000)/12.0
                mdot_ch4 = (data3_mean[-1]*0.665176/60000)/12.0
                phi_corr = 17.136/(mdot_air_tot/mdot_ch4)
                phi_corr_list.append(phi_corr)
                for q in range(len(quant_id)):
                    data_mean, data_uncert, x_plot = self.sweep_zaber(data, t_pretrigger, 'x_pos', quant_id[q], quant_label[q],
                                                 freq_acq_GP)  # 'y_pos:TC, 'x_pos':GP
                    try:
                        value_list[quant_id[q]].append(data_mean[-1])
                    except:
                        value_list[quant_id[q]] = [data_mean[-1]]


                NO_dry,NO2_dry,O2_wet,CO_corr,H2O_conc=self.emission_correction(value_list["NO"][-1], value_list["NO2"][-1], value_list["CO"][-1],
                                         value_list["CO2"][-1], value_list["CH4"][-1], value_list["O2"][-1])
                excess_O2 = self.excess_O2(mdot_ch4,mdot_air_tot,value_list["CO"][-1],value_list["CO2"][-1], value_list["CH4"][-1], value_list["O2"][-1],H2O_conc)
                NOx_corr.append((NO2_dry+NO_dry)*1e6)
                O2_wet_list.append(O2_wet*100)
                CO_corr_list.append(CO_corr*1e6)
                H2O_conc_list.append(H2O_conc)
                excess_O2_list.append(excess_O2)
            ax.scatter(phi_corr_list,NOx_corr, s=20,marker=marker_list[0],color=color_list[count])
            ax2.scatter(phi_corr_list,O2_wet_list, s=20,marker=marker_list[0],color=color_list[count])
            ax3.scatter(phi_corr_list, CO_corr_list, s=20,marker=marker_list[0],color=color_list[count])
            ax4.scatter(phi_corr_list, H2O_conc_list, s=20,marker=marker_list[0],color=color_list[count])
            ax5.scatter(phi_corr_list, excess_O2_list, s=20,marker=marker_list[0],color=color_list[count])

        ax.legend(xlegend, title="kW_Diluent")
        ax.set_ylabel("NOx dry at 15% O2 (ppm)")
        ax.set_xlabel("Equivalence Ratio")
        fig_name = "NOx corrected_Port"+str(port)
        fig.savefig(path+fig_name + '.pdf')
        fig.savefig(path+fig_name + '.png')
        ax2.legend(xlegend, title="kW_Diluent")
        ax2.set_ylabel("O2 wet (vol%)")
        ax2.set_xlabel("Equivalence Ratio")
        fig_name = "O2 wet_Port"+str(port)
        fig2.savefig(path+fig_name + '.pdf')
        fig2.savefig(path+fig_name + '.png')
        ax3.legend(xlegend, title="kW_Diluent")
        ax3.set_ylabel("CO dry at 15% O2 (ppm)")
        ax3.set_xlabel("Equivalence Ratio")
        fig_name = "CO corrected_Port"+str(port)
        fig3.savefig(path+fig_name + '.pdf')
        fig3.savefig(path+fig_name + '.png')
        ax4.legend(xlegend, title="kW_Diluent")
        ax4.set_ylabel("H2O mole fraction")
        ax4.set_xlabel("Equivalence Ratio")
        fig_name = "H2O concentration_Port"+str(port)
        fig4.savefig(path+fig_name + '.pdf')
        fig4.savefig(path+fig_name + '.png')
        ax5.legend(xlegend, title="kW_Diluent")
        ax5.set_ylabel("Excess O2 fraction")
        ax5.set_xlabel("Equivalence Ratio")
        fig_name = "Excess O2_Port" + str(port)
        fig5.savefig(path + fig_name + '.pdf')
        fig5.savefig(path + fig_name + '.png')
        plt.show()

    def correctedemissions_vs_equivalenceratio_fixedpt(self,path):

        t_pretrigger = 0
        quant_id = ['NO', 'NO2', 'CO', 'CO2', 'CH4', 'O2']
        quant_label = ['NO (ppm)', 'NO2 (ppm)', 'CO (ppm)', 'CO2 (vol%)', 'CH4 (ppm)', 'O2 (vol%)']
        extra_quant_id = ['mdot_air', 'mdot_pilotair', 'mdot_ch4']
        extra_quant_label = ['mdot_air', 'mdot_pilotair', 'mdot_ch4']
        freq_acq_GP = 1
        fig, ax = plt.subplots()
        fig2, ax2 = plt.subplots()
        fig3, ax3 = plt.subplots()
        fig4, ax4 = plt.subplots()
        fig5, ax5 = plt.subplots()
        xlegend = []
        marker_list = ["o", "x", "x"]
        color_list = ['k', 'r', 'm', 'g']
        count =-1
        for power in [30]:#,40,50,60]:
            xlegend.append(str(power))
            value_list = {}
            NOx_corr = []
            O2_wet_list=[]
            CO_corr_list=[]
            H2O_conc_list = []
            excess_O2_list=[]
            phi_corr_list=[]
            count+=1
            for phi in self.phi_list:
                dilution = "20.96"
                diluent = "None"
                data = self.dataset[power][phi][dilution][diluent]
                data1_mean, data1_uncert = self.singlept_dataacq_extract(data,extra_quant_id[0])
                data2_mean, data2_uncert = self.singlept_dataacq_extract(data,extra_quant_id[1])
                data3_mean, data3_uncert = self.singlept_dataacq_extract(data,extra_quant_id[2])
                mdot_air_tot = ((data1_mean[-1]+data2_mean[-1])*1.1973/60000)/12.0
                mdot_ch4 = (data3_mean[-1]*0.665176/60000)/12.0
                phi_corr = 17.136/(mdot_air_tot/mdot_ch4)
                phi_corr_list.append(phi_corr)
                for q in range(len(quant_id)):
                    data_mean, data_uncert = self.singlept_dataacq_extract(data, quant_id[q])
                    try:
                        value_list[quant_id[q]].append(data_mean[-1])
                    except:
                        value_list[quant_id[q]] = [data_mean[-1]]


                NO_dry,NO2_dry,O2_wet,CO_corr,H2O_conc=self.emission_correction(value_list["NO"][-1], value_list["NO2"][-1], value_list["CO"][-1],
                                         value_list["CO2"][-1], value_list["CH4"][-1], value_list["O2"][-1])
                excess_O2 = self.excess_O2(mdot_ch4,mdot_air_tot,value_list["CO"][-1],value_list["CO2"][-1], value_list["CH4"][-1], value_list["O2"][-1],H2O_conc)
                NOx_corr.append((abs(NO2_dry)+NO_dry)*1e6)
                O2_wet_list.append(O2_wet*100)
                CO_corr_list.append(CO_corr*1e6)
                H2O_conc_list.append(H2O_conc)
                excess_O2_list.append(excess_O2)
            ax.scatter(phi_corr_list,NOx_corr, s=20,marker=marker_list[0],color=color_list[count])
            ax2.scatter(phi_corr_list,O2_wet_list, s=20,marker=marker_list[0],color=color_list[count])
            ax3.scatter(phi_corr_list, CO_corr_list, s=20,marker=marker_list[0],color=color_list[count])
            ax4.scatter(phi_corr_list, H2O_conc_list, s=20,marker=marker_list[0],color=color_list[count])
            ax5.scatter(phi_corr_list, excess_O2_list, s=20,marker=marker_list[0],color=color_list[count])

        ax.legend(xlegend, title="kW_Diluent")
        ax.set_ylabel("NOx dry at 15% O2 (ppm)")
        ax.set_xlabel("Equivalence Ratio")
        fig_name = "NOx corrected"
        fig.savefig(path+fig_name + '.pdf')
        fig.savefig(path+fig_name + '.png')
        ax2.legend(xlegend, title="kW_Diluent")
        ax2.set_ylabel("O2 wet (vol%)")
        ax2.set_xlabel("Equivalence Ratio")
        fig_name = "O2 wet"
        fig2.savefig(path+fig_name + '.pdf')
        fig2.savefig(path+fig_name + '.png')
        ax3.legend(xlegend, title="kW_Diluent")
        ax3.set_ylabel("CO dry at 15% O2 (ppm)")
        ax3.set_xlabel("Equivalence Ratio")
        fig_name = "CO corrected"
        fig3.savefig(path+fig_name + '.pdf')
        fig3.savefig(path+fig_name + '.png')
        ax4.legend(xlegend, title="kW_Diluent")
        ax4.set_ylabel("H2O mole fraction")
        ax4.set_xlabel("Equivalence Ratio")
        fig_name = "H2O concentration"
        fig4.savefig(path+fig_name + '.pdf')
        fig4.savefig(path+fig_name + '.png')
        ax5.legend(xlegend, title="kW_Diluent")
        ax5.set_ylabel("Excess O2 fraction")
        ax5.set_xlabel("Equivalence Ratio")
        fig_name = "Excess O2"
        fig5.savefig(path + fig_name + '.pdf')
        fig5.savefig(path + fig_name + '.png')
        plt.show()

    def correctedemissions_vs_radialoc(self,path,port):

        t_pretrigger = 30
        quant_id = ['NO', 'NO2', 'CO', 'CO2', 'CH4', 'O2']
        quant_label = ['NO (ppm)', 'NO2 (ppm)', 'CO (ppm)', 'CO2 (vol%)', 'CH4 (ppm)', 'O2 (vol%)']
        extra_quant_id = ['mdot_air', 'mdot_pilotair', 'mdot_ch4']
        extra_quant_label = ['mdot_air', 'mdot_pilotair', 'mdot_ch4']
        freq_acq_GP = 1
        fig, ax = plt.subplots()
        fig2, ax2 = plt.subplots()
        fig3, ax3 = plt.subplots()
        fig4, ax4 = plt.subplots()
        fig5, ax5 = plt.subplots()
        fig6, ax6 = plt.subplots()
        xlegend = []
        marker_list = ["o", "x", "x"]
        color_list = ['k', 'r', 'm', 'g']
        count =-1
        radius_chamb=110#mm
        for power in [30]:#,40,50,60]:
            for phi in self.phi_list:
                value_list = {}
                NOx_corr = []
                O2_wet_list = []
                CO_corr_list = []
                H2O_conc_list = []
                CH4_conc_list = []
                excess_O2_list = []
                phi_corr_list = []
                count += 1
                xlegend.append(str(phi))
                dilution = "20.96"
                diluent = "None"
                data = self.dataset[power][phi][dilution][diluent]
                data1_mean, data1_uncert, x1_plot = self.sweep_zaber(data, t_pretrigger, 'x_pos', extra_quant_id[0],
                                                                     extra_quant_label[0],freq_acq_GP)
                data2_mean, data2_uncert, x2_plot = self.sweep_zaber(data, t_pretrigger, 'x_pos', extra_quant_id[1],
                                                                     extra_quant_label[1],freq_acq_GP)
                data3_mean, data3_uncert, x3_plot = self.sweep_zaber(data, t_pretrigger, 'x_pos', extra_quant_id[2],
                                                                     extra_quant_label[2],freq_acq_GP)
                mdot_air_tot = ((data1_mean[-1]+data2_mean[-1])*1.1973/60000)/12.0
                mdot_ch4 = (data3_mean[-1]*0.665176/60000)/12.0
                phi_corr = 17.136/(mdot_air_tot/mdot_ch4)
                phi_corr_list.append(phi_corr)
                for q in range(len(quant_id)):
                    data_mean, data_uncert, x_plot = self.sweep_zaber(data, t_pretrigger, 'x_pos', quant_id[q], quant_label[q],
                                                 freq_acq_GP)  # 'y_pos:TC, 'x_pos':GP
                    value_list[quant_id[q]] = data_mean
                for x in range(len(x_plot)):
                    NO_dry,NO2_dry,O2_wet,CO_corr,H2O_conc=self.emission_correction(value_list["NO"][x], value_list["NO2"][x], value_list["CO"][x],
                                             value_list["CO2"][x], value_list["CH4"][x], value_list["O2"][x])
                    excess_O2 = self.excess_O2(mdot_ch4,mdot_air_tot,value_list["CO"][x],value_list["CO2"][x], value_list["CH4"][x], value_list["O2"][x],H2O_conc)
                    NOx_corr.append((NO2_dry+NO_dry)*1e6)
                    O2_wet_list.append(O2_wet*100)
                    CO_corr_list.append(CO_corr*1e6)
                    H2O_conc_list.append(H2O_conc)
                    excess_O2_list.append(excess_O2)
                    CH4_conc_list.append(value_list["CH4"][x])
                ax.scatter((x_plot-min(x_plot))/radius_chamb,NOx_corr, s=20,marker=marker_list[0],color=color_list[count])
                ax2.scatter((x_plot-min(x_plot))/radius_chamb,O2_wet_list, s=20,marker=marker_list[0],color=color_list[count])
                ax3.scatter((x_plot-min(x_plot))/radius_chamb, CO_corr_list, s=20,marker=marker_list[0],color=color_list[count])
                ax4.scatter((x_plot-min(x_plot))/radius_chamb, H2O_conc_list, s=20,marker=marker_list[0],color=color_list[count])
                ax5.scatter((x_plot-min(x_plot))/radius_chamb, excess_O2_list, s=20,marker=marker_list[0],color=color_list[count])
                ax6.scatter((x_plot - min(x_plot)) / radius_chamb, CH4_conc_list, s=20, marker=marker_list[0],
                            color=color_list[count])

        ax.legend(xlegend, title="$\phi$")
        ax.set_ylabel("NOx dry at 15% O2 (ppm)")
        ax.set_xlabel("Distance from wall/Radius of Chamber")
        fig_name = "NOx corrected_Port"+str(port)+"_vs_radialloc"
        fig.savefig(path+fig_name + '.pdf')
        fig.savefig(path+fig_name + '.png')
        ax2.legend(xlegend, title="$\phi$")
        ax2.set_ylabel("O2 wet (vol%)")
        ax2.set_xlabel("Distance from wall/Radius of Chamber")
        fig_name = "O2 wet_Port"+str(port)+"_vs_radialloc"
        fig2.savefig(path+fig_name + '.pdf')
        fig2.savefig(path+fig_name + '.png')
        ax3.legend(xlegend, title="$\phi$")
        ax3.set_ylabel("CO dry at 15% O2 (ppm)")
        ax3.set_xlabel("Distance from wall/Radius of Chamber")
        fig_name = "CO corrected_Port"+str(port)+"_vs_radialloc"
        fig3.savefig(path+fig_name + '.pdf')
        fig3.savefig(path+fig_name + '.png')
        ax4.legend(xlegend, title="$\phi$")
        ax4.set_ylabel("H2O mole fraction")
        ax4.set_xlabel("Distance from wall/Radius of Chamber")
        fig_name = "H2O concentration_Port"+str(port)+"_vs_radialloc"
        fig4.savefig(path+fig_name + '.pdf')
        fig4.savefig(path+fig_name + '.png')
        ax5.legend(xlegend, title="$\phi$")
        ax5.set_ylabel("Excess O2 fraction")
        ax5.set_xlabel("Distance from wall/Radius of Chamber")
        fig_name = "Excess O2_Port" + str(port)+"_vs_radialloc"
        fig5.savefig(path + fig_name + '.pdf')
        fig5.savefig(path + fig_name + '.png')
        ax6.legend(xlegend, title="$\phi$")
        ax6.set_ylabel("CH4 (ppm)")
        ax6.set_xlabel("Distance from wall/Radius of Chamber")
        fig_name = "CH4_Port" + str(port) + "_vs_radialloc"
        fig6.savefig(path + fig_name + '.pdf')
        fig6.savefig(path + fig_name + '.png')
        plt.show()

    def extract_data(self, filename,file_preamble,subdir):
        dataextract = DataExtract()
        dataextract.drive = self.drive
        dataextract.folder = self.folder+subdir
        dataextract.quant = self.quant_list

        dataextract.template = lambda x:file_preamble#lambda x:self.file_preamble+str(self.phi_list[x-1])
        dataextract.addon=""
        dataextract.position_range =2
        dataextract.start =1
        nameset = dataextract.drive + dataextract.folder + filename
        try:
            data = dataextract.main()

            with open(dataextract.drive+dataextract.folder + filename, 'wb') as file:
                pickle.dump(data, file, pickle.HIGHEST_PROTOCOL)
        except:
            # cases where combination of diluent, dilution and power do not exist
            data ={}
        return data

    def search_files_temperature(self,power,diluent,dilution,phi,name_data,port):
        directory = self.drive+self.folder
        subdir_list = [x[0] for x in os.walk(directory)]
        for subdir in subdir_list:
            if "Thermocouple_Port"+str(port) in subdir:
                break
        if diluent=="None":
            if "GasComp" in name_data:
                # if files where gas comp and temperature obtained in same file
                filename = "/gasprobe_temp_" + str(power) + "kW_phi" + str(phi)+ ".pkl"
            else:
                # else there is a separate file for it
                filename = "temperature_" + str(power) + "kW_phi" + str(phi) + ".pkl"
        else:
            filename = "/gasprobe_temp_" + str(power) + "kW_phi" + str(phi)+"_" + diluent + "_" + str(dilution) + "perc.pkl"
        template = str(name_data) + str(power) + "kW_phi" + str(phi)
        sub2dir_list = [x[0] for x in os.walk(subdir+"/")]#["O:/GasComposition and Temperature_April2022/GasProbe_Port6_Thermocouple_Port3/Experiments_50kW_20220426/"]#
        sub2dir_list= sub2dir_list[1:-1]
        if len(sub2dir_list)>1:
            for sub2dir in sub2dir_list:
                sub2dir = sub2dir+"/"
                try:
                    with open(sub2dir + filename, 'rb') as f:
                        data = pickle.load(f)
                    return data
                except:

                    data = self.extract_data(filename,template,sub2dir.replace(directory,"",1))

                    if data=={}:
                        continue
                    else:
                        return data
        else:
            subdir =subdir +'/'
            try:
                with open(subdir + filename, 'rb') as f:
                    data = pickle.load(f)
                return data
            except:

                data = self.extract_data(filename, template, subdir.replace(directory, "", 1))
                if port == 6:
                    print(port)
                if data == {}:
                    return []
                else:
                    return data


        return []

    def temperature_unified_dataset(self):
        wb = load_workbook(filename='Case Matrix.xlsx')
        temp_sheet = wb['Temperature measurments']
        self.folder = "GasComposition and Temperature_April2022/"
        cell_y_start = 3
        for i in range(32):
            id = i+cell_y_start
            power = temp_sheet.cell(row =id,column=1).value
            phi = temp_sheet.cell(row =id,column=2).value
            diluent = str(temp_sheet.cell(row =id,column=3).value)
            dilution = temp_sheet.cell(row =id,column=4).value
            Port = temp_sheet.cell(row =id,column=5).value
            name_data = str(temp_sheet.cell(row =id,column=6).value)
            contin_trav = str(temp_sheet.cell(row =id,column=7).value)
            pointwise = str(temp_sheet.cell(row =id,column=8).value)
            zerloc = temp_sheet.cell(row =id,column=9).value
            data = self.search_files_temperature(power,diluent,dilution,phi,name_data,Port)
            #self.dataset_temperature[power] = {phi: {dilution: {diluent: data}}}
            try:
                self.dataset_temperature[Port][pointwise][power][phi][dilution][diluent] = data
            except:
                try:
                    self.dataset_temperature[Port][pointwise][power][phi][dilution] = {diluent: data}
                except:
                    try:
                        self.dataset_temperature[Port][pointwise][power][phi] = {dilution:{diluent: data}}
                    except:
                        try:
                            self.dataset_temperature[Port][pointwise][power] = {phi:{dilution: {diluent: data}}}
                        except:
                            try:
                                self.dataset_temperature[Port][pointwise]={power: {phi: {dilution: {diluent: data}}}}
                            except:
                                self.dataset_temperature[Port]={pointwise:{power: {phi: {dilution: {diluent: data}}}}}

        with open(self.drive+self.folder + "Unified_dataset_temperature", 'wb') as file:
            pickle.dump(self.dataset_temperature, file, pickle.HIGHEST_PROTOCOL)



    def main_temperature(self):
        self.folder = "GasComposition and Temperature_April2022/"
        try:
            with open(self.drive+self.folder + "Unified_dataset_temperature", 'rb') as f:
                self.dataset_temperature = pickle.load(f)
        except:
            self.temperature_unified_dataset()

        port=6
        #self.temperature_vs_equivalenceratio(port=3)
        power = 50
        phi=0.7
        diluent = "None"
        dilution = 20.95
        #self.temperatures_vs_radialoc(port)
        self.temperatures_vs_radialoc_slowtrav_vs_pointwise(port)
        #self.peaktemperatures_vs_equivalenceratio(port)


    def main_gasprobe(self):
        port=3
        if port==6:
            self.folder = self.folder+"GasProbe_Port6_Thermocouple_Port3/"
        else:
            self.folder = self.folder + "GasProbe_Port3/"
        try:
            with open(self.drive+self.folder + "Unified_dataset_gascomp", 'rb') as f:
                self.dataset = pickle.load(f)
        except:
            self.unified_dataset()



        power = 30
        phi = 0.9
        diluent = "CO2"
        dilution = "17.5"
        #self.emissions_vs_radialoc(power,phi,diluent,dilution)
        #self.emissions_vs_equivalenceratio()
        #self.emissions_vs_dilution()
        save_path = self.drive+"GasComposition and Temperature_April2022/Data plots/Gas Comp/"
        #self.correctedemissions_vs_equivalenceratio(save_path,port)
        self.correctedemissions_vs_radialoc(save_path,port)

    def main_exhaust_gasprobe(self):
        self.drive = 'O:/'
        self.folder = "ExhaustModification_June2022/"
        directory = self.drive + self.folder
        self.power_list = [30,60]  # , 40, 50, 60]
        self.phi_list = [0.95, 0.9, 0.8, 0.7, 0.6]#,0.5]
        subdir_list = [x[0]+"/" for x in os.walk(directory)]
        subdir_list = [subdir_list[1]]
        readdata="no"
        if readdata=="yes":
            for subdir in subdir_list:
                for power in self.power_list:
                    self.dataset[power] = {}
                    for phi in self.phi_list:
                        self.dataset[power][phi] = {}
                        filename = "/gasprobe_temp_" + str(power) + "kW_phi" + str(phi) + ".pkl"
                        template = str(power)+"kW_phi"+str(phi)+"_exhausttemp700"
                        data = self.extract_data(filename, template, subdir.replace(directory, "", 1))

        #Unified dataset
        try:
            with open(self.drive+self.folder + "Unified_dataset_gascomp", 'rb') as f:
                self.dataset = pickle.load(f)
        except:
            self.exhaust_unified_dataset()

        save_path = self.drive + "ExhaustModification_June2022/Modified Exhaust/Data plots/"
        self.correctedemissions_vs_equivalenceratio_fixedpt(save_path)









if __name__=="__main__":
    PI = Profiles_intercase()
    #PI.main_exhaust_gasprobe()
    PI.main_gasprobe()
    #PI.main_temperature()
    #path = "O:/GasComposition and Temperature_April2022/GasProbe_Port6_Thermocouple_Port3/Experiments_50kW_20220426/"
    #name=PI.search_tdms_file(path, "Temp_slowtrav_GP6_TC3_50kW_phi0.9")
    #print name